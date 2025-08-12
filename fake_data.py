from faker import Faker
import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Initialize Faker
fake = Faker()

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Contacts"

# Write the header row
headers = ["First Name", "Phone Number", "Company", "Company Size"]
ws.append(headers)

# Generate and add fake data
for _ in range(20):  # 20 fake rows
    first_name = fake.first_name()
    phone_number = fake.msisdn()[:10]  # first 10 digits
    company = fake.company()
    company_size = random.randint(1, 500)
    ws.append([first_name, phone_number, company, company_size])

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# Save the workbook
excel_file = "fake_contacts.xlsx"
wb.save(excel_file)

print(f"✅ Fake contacts saved to '{excel_file}'")
